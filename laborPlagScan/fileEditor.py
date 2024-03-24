import os
import pickle
import tkinter
import zipfile
import tkinter as tk
from tkinter import filedialog
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus.doctemplate import LayoutError
import shutil
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import laborPlagScan.Gui.mainGUI as mainGUI
from laborPlagScan.DataModels.plagiatPaare import PlagiatPaare
from laborPlagScan.DataModels.student import Student


def extract_zip(zip_path):
    """Extracts a zip file to a folder with the same name as the zip file"""
    with zipfile.ZipFile(zip_path, 'r') as zip_file:
        folder_name = os.path.splitext(os.path.basename(zip_path))[0]
        extract_path = os.path.join(os.path.dirname(zip_path), folder_name)
        zip_file.extractall(path=extract_path)
    return extract_path


def unpackZipFiles(folder_path: str):
    """Entpackt alle ZIP-Dateien im Ordner und gibt den Pfad zum Ordner zurück, in dem die Dateien entpackt wurden"""
    # 1. Liste aller Dateien im Ordner
    files = os.listdir(folder_path)
    zipfile_found = False

    # 1.1. Einstellen der Progressbar
    mainGUI.GUI.set_progressbar_start(len(files))

    # 2. Schleife für alle Dateien im Ordner
    for file in files:
        mainGUI.GUI.update_progressbar_value(1)
        # 3. Prüfen, ob Datei eine ZIP-Datei ist
        if not file.endswith('.zip'):
            continue

        zipfile_found = True

        # 4. Entpacken der ZIP-Datei
        file_path = os.path.join(folder_path, file)
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            extract_path = os.path.join(folder_path, "extracted", file.replace(".zip", "").replace(" - ", "_").replace(".java", "").replace(".", "_"))
            zip_file.extractall(path=extract_path)

    # 5. Wenn keine ZIP-Datei gefunden wurde, nach & in Unterordnern suchen
    if not zipfile_found:
        subfolders = os.listdir(folder_path)
        for subfolder in subfolders:
            subfolder_path = os.path.join(folder_path, subfolder)
            if not os.path.isdir(subfolder_path):
                continue
            extract_path = unpackZipFiles(subfolder_path)
            return extract_path

    return os.path.dirname(extract_path)


def remove_folder(folder_path):
    """Removes a folder and all its content"""
    shutil.rmtree(folder_path)


def get_last_modified_file():
    """Returns the last modified file in the result folder"""
    folder_path = 'laborPlagScan/result'
    if not os.path.exists(folder_path):
        return None
    files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    if not files:
        return None
    last_modified_file = max(files, key=os.path.getmtime)
    last_modified_file = os.path.basename(last_modified_file).split(".")[0]
    return last_modified_file


def save_auswertung_to_file(obj, path="laborPlagScan/result/", filename="last_result"):
    """Saves the auswertung Objects to a pickle-file"""
    if path is None:
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askdirectory()
        root.destroy()

    filename = filename.split('.')[0] + '.pkl'
    path = os.path.join(path, filename)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'wb') as outp:  # Schreibt die Datei im Binärmodus
        pickle.dump(obj, outp, pickle.HIGHEST_PROTOCOL)


def load_auswertung_from_file(filename=None) -> [[PlagiatPaare], [Student]]:
    """Loads the 'auswertung' Objects from a pickle-file"""
    if filename is None:
        filepath = tkinter.filedialog.askopenfilename(initialdir="/", title="Select file")
        FileNotFoundError("No file selected")

    else:
        # Erstelle den Pfad zur Datei im Ordner "result"
        filepath = f"laborPlagScan/result/{filename}.pkl"

    # Öffne die Datei und lese die Objekte aus
    with open(filepath, 'rb') as inp:  # Liest die Datei im Binärmodus
        return pickle.load(inp)


def get_string_lines(string: str, lines: list):
    """Returns the first 'lines' lines of the string"""

    content = string.split("\n")
    for i in range(len(content)):
        content[i] += "\n"

    content_parts = []
    start, end = lines
    # before
    for i in content[max(start - 5, 0):max(start - 1, 0)]:
        content_parts.append(i)
    content_parts.append("\n\t\t------PLAGIAT START------\n\n")

    # plagiat
    for i in content[max(start - 1, 0):end]:
        content_parts.append(i)

    # after
    content_parts.append("\n\t\t-------PLAGIAT END-------\n\n")
    for i in content[end:min(end + 5, len(content))]:
        content_parts.append(i)

    return ''.join(content_parts)


def table_to_pdf_export(savepath: str, data: list, format=A4):

    temp_filename = 'laborPlagScan/result/temp.pdf'

    try:
        table = Table(data)

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ]))

        doc = SimpleDocTemplate(temp_filename, pagesize=format, rightMargin=10, leftMargin=10, topMargin=5, bottomMargin=5)

        elements = [table]
        doc.build(elements)

        shutil.move(temp_filename, savepath)

        print(f"Die PDF-Datei '{savepath}' wurde erfolgreich im Format {'Querformat' if format == landscape(A4) else 'Hochformat'} erstellt.")
    except LayoutError as e:
        if format == landscape((297, 210)):
            raise Exception("LayoutError konnte nicht automatisch behoben werden.")

        print("LayoutError erkannt, wechsle zu Querformat und versuche erneut.")
        table_to_pdf_export(savepath, data, format=landscape((297, 210)))


def table_to_xlsx_export(savepath: str, data: list):
    """
    Exports the data to a xlsx file
    :param savepath: savepath of the xlsx file as string
    :param data: Tabledata as list of lists
    """

    def get_table_width(table_data):
        """Returns the width of the table in cm"""
        width_multiplyer: float = 0.83
        height_multiplyer: float = 14.8
        table_width = []
        table_height = [0 for pos in range(len(table_data))]
        for data_column, col in enumerate(table_data[0]):
            column_width = 0
            column_width_sum, column_width_summanden = 0, 0
            for row_nr, data_row in enumerate(table_data):
                data_cell_inhalt = str(data_row[data_column]).split("\n")
                table_height[row_nr] = max(table_height[row_nr], len(data_cell_inhalt))
                for inhalt in data_cell_inhalt:
                    column_width_sum += len(inhalt)
                    column_width_summanden += 1
                    column_width = max(column_width, len(inhalt))
            column_width = ((column_width_sum / column_width_summanden) + column_width) / 2
            table_width.append(column_width)

        for height_pos in range(len(table_height)):
            table_height[height_pos] = max(16, table_height[height_pos] * height_multiplyer)

        for width_pos in range(len(table_width)):
            table_width[width_pos] = max(16, table_width[width_pos] * width_multiplyer)

        return [table_width, table_height]

    # tablesize = []
    # [Spalt-BreiteA, Spalt-BreiteB, Spalt-BreiteC, ...]
    # [Row-Height1, Row-Height2, Row-Height3, ...]
    tablesize = get_table_width(data)
    workbook = Workbook()
    sheet = workbook.active

    # Daten in das Tabellenblatt einfügen
    row_count = 1
    for row_data in data:
        row = []
        for cell_data in row_data:
            column_count = len(row) + 1
            cell = sheet.cell(row=row_count, column=column_count)
            cell.value = cell_data
            cell.alignment = cell.alignment.copy(wrapText=True)  # Umbrüche beibehalten
            row.append(cell)
        row_count += 1

    # Spaltenbreite anpassen
    for col_nr, column in enumerate(sheet.columns):
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = tablesize[0][col_nr]

    for i in range(0, sheet.max_row):
        sheet.row_dimensions[i + 1].height = tablesize[1][i]

    workbook.save(savepath)
